# %%
import numpy as np
from matplotlib import pyplot as plt
import os
import random
import re
import pandas as pd
from pybnesian import LinearCorrelation, MutualInformation, KMutualInformation, RCoT
import re
import ot 
from scipy.stats import spearmanr
from scipy.stats import pearsonr
from bicausal.helpers.extra import hsic, XtendedCorrel
from bicausal.helpers.namemap import name_map
import openpyxl
import json
from typing import List, Optional

# %%
seedR = random.Random(42)
seedN = np.random.default_rng()

# %%
#AUX:

#Only contains the two variables
#each pair is of size (N,2)
def getnumpy(df):
    data=df.to_numpy()
    tables=[]
    for i in range(data.shape[0]):
        tables.append(np.array([data[i][0], data[i][1]]))
        
    return tables

def getTuebingen(read_dir="benchmarks/Tuebingen"):
    """
    Reads the Tübingen dataset pairs and their weights from the given directory.
    Returns:
        data: list of (x_matrix, y_matrix)
        weights: np.array of weights
    """
    pairmeta_file = os.path.join(read_dir, "pairmeta.txt")
    pair_prefix = "pair"
    data, weights = [], []

    with open(pairmeta_file, "r") as f:
        meta_lines = f.readlines()

    for line in meta_lines:
        entries = line.split()
        pair_number = entries[0].zfill(4)
        x_start, x_end = int(entries[1]) - 1, int(entries[2])
        y_start, y_end = int(entries[3]) - 1, int(entries[4])
        weight = float(entries[5])

        pair_filename = os.path.join(read_dir, f"{pair_prefix}{pair_number}.txt")
        try:
            arr = np.loadtxt(pair_filename)
            x, y = arr[:, x_start:x_end], arr[:, y_start:y_end]
            data.append((x, y))
            weights.append(weight)
        except FileNotFoundError:
            print(f"⚠️ Missing {pair_filename}, skipping.")

    return data, np.array(weights)

# Example usage:
# data, weights = load_pairs_and_weights("pairmeta.txt")


def reorder_by_abs(scores, weights):
    valid_mask = ~np.isnan(scores)
    
    # Filter out NaN entries from both scores and weights
    valid_scores = scores[valid_mask]
    valid_weights = weights[valid_mask]
    # Get indices that would sort the array by descending absolute value
    sorted_indices = np.argsort(-np.abs(valid_scores))
    
    # Apply the sorted indices to both scores and weights
    sorted_scores = valid_scores[sorted_indices]
    sorted_weights = valid_weights[sorted_indices]

    return sorted_scores, sorted_weights

def minmax_scale(data):
    # Compute the minimum and maximum for each column
    min_val = np.nanmin(data, axis=0)
    max_val = np.nanmax(data, axis=0)
    scale = max_val - min_val
    
    # To avoid division by zero, replace any 0 in scale with 1 temporarily
    scale_safe = np.where(scale == 0, 1, scale)
    
    # Perform the min-max scaling
    scaled_data = (data - min_val) / scale_safe
    
    # For columns with constant values, set the entire column to 0
    scaled_data[:, scale == 0] = 0
    return scaled_data

def sign_to_binary(x):
    return np.where(x > 0, 1, np.where(x < 0, 0, 0.5))

def switch_signs(vector):
    # Generate a binary vector (mask) with 0's and 1's equally likely.
    mask = np.random.choice([0, 1], size=vector.shape)
    
    # Make a copy of the original vector to avoid modifying it directly.
    new_vector = vector.copy()
    
    # Switch sign where mask is 0
    new_vector[mask == 0] *= -1
    
    return new_vector, mask

def sanitize_filename(filename):
    # Allow letters, numbers, underscore, dot, dash, and slash
    sanitized = re.sub(r'[^a-zA-Z0-9_.\/-]', '_', filename)
    # Collapse multiple underscores into one
    sanitized = re.sub(r'_+', '_', sanitized)
    # Remove leading/trailing underscores
    return sanitized.strip('_')

def save_imgs(filename, dir):
    os.makedirs(dir, exist_ok=True)  # Create directory if it doesn't exist
    # Ensure a valid filename
    filename = f"{dir}/{sanitize_filename(filename)}.png"
    plt.savefig(filename, dpi=300, bbox_inches='tight')

def save_table(df,title,filename, dir):
    df = df.round(3)
    fig_width = df.shape[1] * 1.2
    fig_height = df.shape[0] * 0.8
    fig, ax = plt.subplots(figsize=(fig_width, fig_height))
    ax.axis('tight')
    ax.axis('off')
    plt.title(title)
    # Create the table
    table = ax.table(cellText=df.values, colLabels=df.columns, rowLabels=df.index, cellLoc='center', loc='center')
    filename = f"{dir}/{sanitize_filename(filename)}.png"
    # Save the table as an image
    plt.savefig(filename, dpi=300, bbox_inches='tight')

def wasserstein_distance(x,y):
    M = ot.dist(x, y, metric='euclidean')
    n_samples=len(x)
    a = np.ones(n_samples) / n_samples
    b = np.ones(n_samples) / n_samples
    wasserstein_distance_squared = ot.emd2(a, b, M)
    return np.sqrt(wasserstein_distance_squared)

def remove_outliers(x, y,per=0.9):
    x, y = np.array(x), np.array(y)

    # Compute IQR for x and y
    Q1_x, Q3_x = np.percentile(x, [100*(1-per)/2, 100*(per+(1-per)/2)])
    #IQR_x = Q3_x - Q1_x
    #lower_x, upper_x = Q1_x - mul * IQR_x, Q3_x + mul* IQR_x
    lower_x,upper_x=Q1_x,Q3_x

    Q1_y, Q3_y = np.percentile(y, [100*(1-per)/2, 100*(per+(1-per)/2)])
    #IQR_y = Q3_y - Q1_y
    #lower_y, upper_y = Q1_y - mul * IQR_y, Q3_y + mul * IQR_y
    lower_y,upper_y=Q1_y,Q3_y

    # Create a mask for valid (non-outlier) points
    mask = (x >= lower_x) & (x <= upper_x) & (y >= lower_y) & (y <= upper_y)
    return x[mask], y[mask]

def derivative_polyfit(time, samples, r, poly_degree=2):
    """
    Estimate the derivative by fitting a polynomial to the last r samples and 
    differentiating the fitted polynomial at the most recent time.

    Parameters:
        time (list or np.array): Timestamps corresponding to each sample in chronological order,
                                 with the most recent sample at the end.
        samples (list or np.array): Data points corresponding to the time points.
        r (int): The number of recent samples to use for the fit.
        poly_degree (int): Degree of the polynomial to fit (default is 2).

    Returns:
        float: The estimated derivative at the current time.

    Raises:
        ValueError: If the number of samples or time points is less than r, 
                    or if the lengths of time and samples do not match.
    """
    if len(samples) < r or len(time) < r:
        raise ValueError(f"At least {r} samples are required.")
    if len(time) != len(samples):
        raise ValueError("Time and samples vectors must have the same length.")
    
    # Use the last r samples.
    y = np.array(samples[-r:])
    t_recent = np.array(time[-r:])
    
    # Shift the time vector so that the most recent sample corresponds to t = 0.
    t_shifted = t_recent - t_recent[-1]
    
    # Fit a polynomial of the given degree to the (t_shifted, y) data.
    coeffs = np.polyfit(t_shifted, y, poly_degree)
    
    # Compute the derivative of the polynomial.
    poly_deriv = np.polyder(coeffs)
    
    # Evaluate the derivative at t = 0 (the most recent sample).
    deriv_at_zero = np.polyval(poly_deriv, 0)
    
    return deriv_at_zero

# %%
#Completely random
def random_fun(data):
    if seedN.random()<0.2:
        return np.nan
    return seedR.random()-0.5


def save_excel(funcname, dataset, metrics, metric_titles, **kwargs):
    """
    Executes func with the provided kwargs, computes metrics from the result,
    and appends or overwrites a row in an Excel file ("results.xlsx") where:
      - The first column is the method name (),
      - The second column is the dataset,
      - The third column is a string of all kwargs,
      - Subsequent columns are the computed metric values.
      
    Parameters:
        func (function): The function to run.
        dataset (any): Identifier for the dataset.
        metrics (list of functions): Functions that compute a metric from func's output.
        metric_titles (list of str): Column titles for the metric values.
        **kwargs: Additional keyword arguments to pass to func.
    """
    
    # Convert kwargs to a readable string format
    kwargs_str = ", ".join(f"{k}={v}" for k, v in kwargs.items())
    
    # Prepare the row: [Method, Dataset, Parameters, ...metrics]
    row = [funcname, dataset, kwargs_str] + metrics

    file_path = "../previous_results.xlsx"
    if os.path.exists(file_path):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        # Write header row if file is new
        header = ["Method", "Dataset", "Parameters"] + metric_titles
        ws.append(header)
    
    # Check if a row with the same Method, Dataset, and Parameters exists
    for row_idx, existing_row in enumerate(ws.iter_rows(values_only=True), start=1):
        #print(existing_row[2], kwargs_str, "==", (existing_row[2] or "")  == kwargs_str)
        if (existing_row[0] == funcname and 
            existing_row[1] == dataset and 
            (existing_row[2] or "") == kwargs_str):
            # Overwrite the existing row
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
            wb.save(file_path)
            return
    
    # Append the new row if no match is found
    ws.append(row)
    wb.save(file_path)


# %%
def test_independence(x,y,method="pearson", **kwargs):
    x=x.flatten()
    y=y.flatten()
    df = pd.DataFrame({"x": x, "y": y})

    if method=="pearson":
        pearson_corr, pearson_p = pearsonr(x, y)
        return - pearson_p
    elif method=="spearman":
        spearman_corr, spearman_p = spearmanr(x, y)
        return - spearman_p
    elif method=="hsic":
        return hsic.dHSIC(x, y) / np.sqrt(hsic.dHSIC(x, x) * hsic.dHSIC(y, y))
    elif method=="hoef":
        return XtendedCorrel.hoeffding(x,y)
    """
            The independence score. Note:
                - For "linear" and "rcot": a higher p-value means more evidence of independence.
                - For "mutual_info" and "kmutual": a score of 0 means independence, and higher values indicate dependence.
    """
    # Create a DataFrame with column names 'x' and 'y'
    
    
    if method == "linear":
        # For continuous data, use the partial linear correlation test.
        test = LinearCorrelation(df)
        # pvalue: higher p-value => cannot reject independence
        score = test.pvalue("x", "y")
        return - score

    elif method == "mutual_info":
        # Using MutualInformation (assumes a Gaussian model for continuous data)
        test = MutualInformation(df, asymptotic_df=True)
        # MI: 0 indicates independence; higher MI means more dependence.
        score = test.mi("x", "y")
        return score

    elif method == "kmutual":
        # Using KMutualInformation, a k-nearest neighbors estimator.
        k = kwargs.get("k", 5)
        seed = kwargs.get("seed", None)
        shuffle_neighbors = kwargs.get("shuffle_neighbors", 5)
        samples = kwargs.get("samples", 1000)
        test = KMutualInformation(df, k, seed=seed, shuffle_neighbors=shuffle_neighbors, samples=samples)
        score = test.mi("x", "y")
        return score

    elif method == "rcot":
        # Using RCoT, a randomized conditional correlation test based on random Fourier features.
        random_fourier_xy = kwargs.get("random_fourier_xy",15)
        random_fourier_z = kwargs.get("random_fourier_z", 100)
        test = RCoT(df, random_fourier_xy=random_fourier_xy, random_fourier_z=random_fourier_z)
        # pvalue: higher p-value indicates more independence.
        score = test.pvalue("x", "y")
        return - score
    elif method=="wasser": #mede o quao longe eq ta de ser ao calhas
        y_perm=np.random.permutation(y)
        dist_true=np.array([x,y])
        dist_perm=np.array([x,y_perm])
        return - wasserstein_distance(dist_true,dist_perm)

def serialize_params(args, kwargs):
    """Serialize only kwargs as compact JSON string for CSV storage."""
    if not kwargs:
        return ""
    try:
        return json.dumps(kwargs, sort_keys=True, separators=(",", ": "))
    except Exception:
        # Fallback: readable string
        return str(kwargs)


def normalize_str(value):
    """
    Normalizes values read from CSV so that NaN or None become ''.
    Ensures empty parameters stay consistent between write and read.
    """
    if pd.isna(value) or value is None:
        return ""
    return str(value) 

def correct_names(folder: str = "results"):

    # Ensure folder exists
    if not os.path.isdir(folder):
        raise FileNotFoundError(f"The folder '{folder}' does not exist.")

    for filename in os.listdir(folder):
        if filename.lower().endswith(".csv"):
            filepath = os.path.join(folder, filename)

            # Load CSV
            df = pd.read_csv(filepath)

            # If "method" column exists, update names
            if "method" in df.columns:
                # Replace only names that exist in the mapping
                df["method"] = df["method"].map(
                    lambda x: name_map[x.lower()] if x.lower() in name_map else x
                )

            # Save back to CSV
            df.to_csv(filepath, index=False)


import os
import pandas as pd

def nanning(dir="results"):
    if not os.path.isdir(dir):
        raise ValueError(f"Directory '{dir}' does not exist.")

    for filename in os.listdir(dir):
        if filename.lower().endswith(".csv"):
            filepath = os.path.join(dir, filename)
            print("Processing", filepath)
            
            # Load CSV normally
            df = pd.read_csv(filepath)
            
            # Check if score column exists
            if df.shape[1] >= 4:  # fourth column exists
                col_name = "score"
                if col_name in df.columns:
                    # Track changes
                    original_values = df[col_name].copy()
                    
                    # Replace empty or NaN or 0 with "NA"
                    df[col_name] = df[col_name].apply(lambda x: "NA" if pd.isna(x) or x == "" or x == 0 else x)
                    # Save back
                    df.to_csv(filepath, index=False)

KNOWN_TIMESTAMP_FORMATS: List[str] = [
    '%Y-%m-%dT%H:%M:%SZ',  # ISO 8601 (e.g., 2025-12-05T23:18:22Z)
    '%Y-%m-%d %H:%M:%S',   # SQL format (e.g., 2025-12-07 21:23:30)
    # Add any other known formats here if they appear
]

def delete_repeated_lines(folder: str = "results") -> None:
    """
    Ultimate fix: Handles mixed timestamp formats, floating point inconsistencies,
    and missing data to robustly find and delete older duplicates.
    """
    if not os.path.isdir(folder):
        raise FileNotFoundError(f"The folder '{folder}' does not exist.")

    for filename in os.listdir(folder):
        if filename.lower().endswith(".csv"):
            filepath = os.path.join(folder, filename)
            
            print(f"--- Processing {filename} ---")
            df = pd.read_csv(filepath)

            if "timestamp" not in df.columns:
                print(f"File {filename} skipped: Missing 'timestamp' column.")
                continue

            # 1. ROBUST TIMESTAMP PARSING (The key change)
            print("Attempting robust timestamp parsing...")
            
            # Start with a column of NaT
            df['parsed_timestamp'] = pd.NaT 
            
            # Iterate through all known formats to try and parse the column
            for fmt in KNOWN_TIMESTAMP_FORMATS:
                # Only try to parse the timestamps that haven't been successfully parsed yet (are still NaT)
                unparsed_mask = df['parsed_timestamp'].isna()
                
                # Use apply on the subset of unparsed rows for targeted conversion
                df.loc[unparsed_mask, 'parsed_timestamp'] = pd.to_datetime(
                    df.loc[unparsed_mask, 'timestamp'], format=fmt, errors='coerce'
                )

            # Update the original column with the successfully parsed values
            df['timestamp'] = df['parsed_timestamp']
            df.drop(columns=['parsed_timestamp'], inplace=True)
            
            # Now, drop the rows that *still* couldn't be parsed after trying all formats
            original_size = len(df)
            df.dropna(subset=['timestamp'], inplace=True)
            dropped_invalid = original_size - len(df)

            if dropped_invalid > 0:
                print(f"Final warning: Dropped {dropped_invalid} rows with completely unparsable timestamps.")
            
            if df.empty:
                print(f"File {filename} empty after cleaning. Skipping save.")
                continue

            # 2. Define columns to exclude and check (Same as before)
            columns_to_exclude = {"score", "timestamp"}
            duplicate_check_cols = [
                col for col in df.columns if col not in columns_to_exclude
            ]

            # 3. Robustness for Float and Missing Data (Same as before)
            for col in duplicate_check_cols:
                # Fill NaN with a distinct string ('MISSING') to treat blanks as equal
                df[col].fillna('MISSING', inplace=True) 
                
                if pd.api.types.is_float_dtype(df[col]):
                    # Round floats to 15 decimal places for stable comparison
                    df[col] = df[col].round(15)

            # 4. Sort and Drop Duplicates
            df.sort_values(by='timestamp', ascending=False, inplace=True)
            df_cleaned = df.drop_duplicates(
                subset=duplicate_check_cols, 
                keep='first'
            )
            
            # 5. Report and Save
            rows_deleted = len(df) - len(df_cleaned)
            if rows_deleted > 0:
                print(f"Successfully deleted {rows_deleted} repeated (older) lines from {filename}.")
                df_cleaned.to_csv(filepath, index=False)
            else:
                print(f"No new repeated lines found in {filename}.")

    print("\n--- Ultimate Deduplication complete. ---")

def tuebingen_statistics(xlsx_path="benchmarks/Tuebingen/TuebingenAnalysis.xlsx"):
    df = pd.read_excel(xlsx_path)

    if not {"Datapoints", "weights"}.issubset(df.columns):
        raise ValueError("Excel must contain columns: datapoints, weight")

    # --- examples per number of points ---
    datapoints = df["Datapoints"].dropna().astype(int).tolist()

    # --- reconstruct number of examples per source ---
    # Weight = 1/k => k = 1/weight
    # Multiple rows with the same weight belong to the same source(s).
    examples_per_source = []

    # Count how many rows have each distinct weight
    weight_counts = df["weights"].dropna().value_counts().to_dict()

    for w, count in weight_counts.items():
        if w <= 0:
            continue
        k = int(round(1.0 / float(w)))  # examples per source for this weight
        if k <= 0:
            continue

        # number of complete sources represented by these rows
        num_full_sources = count // k
        for _ in range(num_full_sources):
            examples_per_source.append(k)

        remainder = count % k

    return datapoints, examples_per_source

import pandas as pd
import numpy as np
from typing import Union

def flip_scores_by_method(scores_path: str, target_method: str) -> pd.DataFrame:
    """
    Loads a CSV file of scores, flips the 'score' column (s = -s) for all
    rows where the 'method' column matches the target_method. NA scores are
    preserved.

    Args:
        scores_path: The file path to the CSV containing the score data.
        target_method: The value in the 'method' column to filter and flip scores for.

    Returns:
        A pandas DataFrame with the modified scores.
    """
    try:
        # Load the CSV. keep_default_na=False prevents pandas from interpreting
        # strings like 'NA' as NaN initially.
        df = pd.read_csv(scores_path, keep_default_na=False)
    except FileNotFoundError:
        print(f"Error: The file '{scores_path}' was not found.")
        return pd.DataFrame()
    except Exception as e:
        print(f"An error occurred while loading the CSV: {e}")
        return pd.DataFrame()

    # 1. Standardize NA values (as requested)
    # Replace the string "NA" with the numpy NaN value (for proper handling)
    df = df.replace("NA", np.nan)

    # 2. Prepare 'parameters' column (as requested)
    # Fill actual NA/NaN with an empty string and ensure it's a string type
    df["parameters"] = df["parameters"].fillna("").astype(str)

    # Convert 'score' to numeric, coercing any non-numeric value that resulted 
    # from the "NA" replacement (now np.nan) to NaN.
    # The 'score' column must be numeric for the negation operation.
    df['score'] = pd.to_numeric(df['score'], errors='coerce')


    # 3. Apply the conditional score flip
    
    # Create a boolean mask where the 'method' matches the target
    condition = df['method'] == target_method
    
    # Select the scores that meet the condition and are NOT NaN (because we
    # must maintain NA scores)
    scores_to_flip = df.loc[condition & df['score'].notna(), 'score']
    
    # Flip the selected scores in place (s = -s)
    df.loc[condition & df['score'].notna(), 'score'] = -scores_to_flip
    
    # Alternative one-liner for the flip using np.where (slightly cleaner)
    # df['score'] = np.where(
    #     (df['method'] == target_method) & df['score'].notna(), 
    #     -df['score'], 
    #     df['score']
    # )

    return df